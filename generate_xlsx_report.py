#!/usr/bin/env python3
"""Generate a professional XLSX report from all Facebook leads (105+)."""
import csv, re
from datetime import datetime
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(['uv', 'pip', 'install', 'openpyxl'], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SRC = '/home/skarbolt/kb/lead-gen/collected_leads/facebook_leads.csv'
OUT = '/home/skarbolt/kb/lead-gen/facebook_leads_report.xlsx'

# ── Load all leads from enriched CSV ──
all_leads = []
with open(SRC, 'r') as f:
    reader = csv.DictReader(f)
    for r in reader:
        r['_name'] = (r.get('name') or '').strip()
        r['_url'] = (r.get('url') or '').strip()
        r['_phone'] = (r.get('phone') or '').strip()
        r['_has_web'] = r.get('has_website', '').strip().lower() in ('yes', 'true', 'y', '1')
        r['_website'] = (r.get('website') or '').strip()
        r['_email'] = (r.get('email') or '').strip()
        r['_notes'] = (r.get('notes') or '').strip()
        r['_ad_status'] = (r.get('ad_status') or '').strip()
        r['_date'] = r.get('date', '')[:10]
        
        try:
            r['_score'] = int(r.get('score', '0') or '0')
        except:
            r['_score'] = 0
        
        try:
            likes_raw = r.get('likes', '0') or '0'
            r['_likes'] = int(re.sub(r'[^0-9]', '', str(likes_raw))) if str(likes_raw).strip() else 0
        except:
            r['_likes'] = 0
        
        if r['_name'] and r['_url']:
            all_leads.append(r)

print(f"Loaded {len(all_leads)} leads total")

# ── Classifications ──
def classify_lead(l):
    """Auto-classify a lead into a category based on name."""
    name = l['_name'].lower()
    if any(w in name for w in ['salon', 'makeover', 'makeup', 'beauty', 'cosmetic', 'skincare', 
                                'skin care', 'nail', 'hair', 'beautyology', 'glam',
                                'fency skin', 'mehedi enterpris', 'lavender studio']):
        return 'Beauty & Cosmetics'
    if any(w in name for w in ['fashion', 'clothing', 'dress', 'wear', 'apparel', 'fabric',
                                'glamgrl', 'easy fashion', 'easy fashion ltd', 'everywear',
                                'rm fashion', 'najar fashion', 'fashion hub', 'fashion brand',
                                'andeem', 'tote']):
        return 'Fashion & Apparel'
    if any(w in name for w in ['baby', 'kids', 'toy', 'moms', 'mum', 'little angel',
                                'child', 'infant', 'kids paradise']):
        return 'Baby & Kids'
    if any(w in name for w in ['electronic', 'gadget', 'mobile', 'phone', 'computer', 
                                'tech', 'laptop', 'brand electronic', 'dubai electronic',
                                'electronic-bd', 'tesla', 'popular electronic']):
        return 'Electronics & Gadgets'
    if any(w in name for w in ['grocery', 'foodi', 'mart', 'supermarket', 'food',
                                'unimart', 'lavender super', 'khulshi', 'halal']):
        return 'Grocery & Food'
    if any(w in name for w in ['handicraft', 'handmade', 'craft', 'hlwcraft', 'palli',
                                'sarmin', 'craftimation', 'kmt']):
        return 'Handicraft & Artisan'
    if any(w in name for w in ['jewel', 'jewellery', 'gold', 'diamond', 'pearl', 'ornament',
                                'royal malabar', 'new dhaka jewel']):
        return 'Jewelry'
    if any(w in name for w in ['home', 'decor', 'interior', 'furniture', 'mirror', 'glass',
                                'aniq', 'naem', 'sohan', 'art work']):
        return 'Home & Decor'
    if any(w in name for w in ['shoe', 'footwear', 'leather', 'walk', 'sparrow',
                                'walkon']):
        return 'Footwear'
    if any(w in name for w in ['gift', 'present', 'bondhon', 'safa gift']):
        return 'Gift Shop'
    if any(w in name for w in ['perfume', 'fragrance', 'elate', 'napa']):
        return 'Perfume & Fragrance'
    if any(w in name for w in ['organic', 'herb', 'rue', 'miss organic']):
        return 'Organic & Health'
    if any(w in name for w in ['sport', 'gym', 'fitness', 'ca sport', 'dugout']):
        return 'Sports & Fitness'
    if any(w in name for w in ['bag', 'accessorie', 'don sumdany', 'story brand',
                                'tijara', 'wide zone', 'jfm']):
        return 'Accessories & Lifestyle'
    return 'General Retail'

def is_running_ads(l):
    status = l['_ad_status'].lower()
    return 'currently running' in status

def is_hot(l):
    return not l['_has_web'] and bool(l['_phone'])

# ── Categorize ──
for l in all_leads:
    l['_category'] = classify_lead(l)

# ── Create workbook ──
wb = Workbook()

# ── Styles ──
NAVY = '1F4E79'
GOLD = 'FFD700'
LIGHT_NAVY = 'D6E4F0'
LIGHT_GREEN = 'E2EFDA'
LIGHT_RED = 'FCE4EC'
WHITE = 'FFFFFF'

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=11)
GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type='solid')
ALT_FILL = PatternFill(start_color=LIGHT_NAVY, end_color=LIGHT_NAVY, fill_type='solid')
GREEN_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
RED_FILL = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')

TITLE_FONT = Font(name='Calibri', bold=True, size=16, color=NAVY)
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color=NAVY)
BOLD = Font(name='Calibri', bold=True, size=10)
NORMAL = Font(name='Calibri', size=10)
RED_BOLD = Font(name='Calibri', bold=True, size=10, color='B22222')
GREEN_BOLD = Font(name='Calibri', bold=True, size=10, color='006100')

THIN_BORDER = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0'),
)

def apply_header(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN_BORDER

def apply_row(ws, row_idx, values, styles=None):
    """Apply values and styles to a row. styles is a list of (col_idx, font, fill) tuples."""
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=v)
        c.font = NORMAL
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Apply custom style if provided
        if styles:
            for sci, sf, sfill in styles:
                if ci == sci:
                    c.font = sf
                    if sfill:
                        c.fill = sfill
        
        # Alternating row color
        if not styles and row_idx % 2 == 0:
            c.fill = ALT_FILL

def auto_width(ws, headers, data_rows, max_samples=50):
    for ci in range(1, len(headers) + 1):
        best = len(str(headers[ci-1]))
        for ri in range(2, min(len(data_rows) + 2, max_samples + 2)):
            cell_val = str(ws.cell(row=ri, column=ci).value or '')
            best = max(best, min(len(cell_val), 55))
        ws.column_dimensions[get_column_letter(ci)].width = best + 3

def write_sheet(ws, headers, data_rows, value_fn, style_fn=None):
    """Write a data sheet. value_fn(row) returns list of values. style_fn(row) returns optional styles list."""
    apply_header(ws, headers)
    for ri, row in enumerate(data_rows, 2):
        vals = value_fn(row)
        styles = style_fn(row) if style_fn else None
        apply_row(ws, ri, vals, styles)
    auto_width(ws, headers, data_rows)
    ws.freeze_panes = 'A2'
    if data_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data_rows)+1}"

# ── Helper functions for row values ──
def main_vals(l):
    return [
        '',  # will be numbered
        l['_name'],
        l['_category'],
        l['_phone'] if l['_phone'] else 'N/A',
        'No' if l['_has_web'] else 'Yes',
        l['_website'] if l['_website'] else '',
        l['_email'] if l['_email'] else '',
        l['_score'],
        f"👍 {l['_likes']:,}" if l['_likes'] else '',
        '🚀 Active' if is_running_ads(l) else '',
        l['_date'],
        l['_url'],
    ]

def main_style(l):
    styles = []
    if is_hot(l):
        for ci in range(1, 13):
            styles.append((ci, RED_BOLD, GOLD_FILL))
    if l['_has_web']:
        pass  # use default
    return styles

# ════════════════════════════════════
# SHEET 1: All Leads
# ════════════════════════════════════
ws1 = wb.active
ws1.title = "All Leads"

H1 = ['#', 'Business Name', 'Category', 'Phone', 'No Website?', 'Website URL', 'Email', 
      'Score', 'Likes', 'Boost Status', 'Date Added', 'Facebook URL']

sorted_leads = sorted(all_leads, key=lambda l: (-l['_score'], -l['_likes'], l['_name']))

# Add row numbers
for i, l in enumerate(sorted_leads):
    l['_num'] = i + 1

write_sheet(ws1, H1, sorted_leads, 
    lambda l: [l['_num']] + main_vals(l)[1:],
    main_style)

# ════════════════════════════════════
# SHEET 2: Active Boosters
# ════════════════════════════════════
ws2 = wb.create_sheet("Active Boosters")
booster_leads = [l for l in sorted_leads if is_running_ads(l)]

write_sheet(ws2, H1, booster_leads,
    lambda l: [booster_leads.index(l)+1] + main_vals(l)[1:],
    main_style)

# ════════════════════════════════════
# SHEET 3: Hot Leads (No Website)
# ════════════════════════════════════
ws3 = wb.create_sheet("Hot Leads (No Website)")
hot_leads = [l for l in sorted_leads if is_hot(l)]

write_sheet(ws3, H1, hot_leads,
    lambda l: [hot_leads.index(l)+1] + main_vals(l)[1:],
    main_style)

# ════════════════════════════════════
# SHEET 4: Dashboard
# ════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")

# Title
ws4.merge_cells('A1:F1')
c = ws4['A1']
c.value = "📊 FACEBOOK LEADS DASHBOARD — SKARBOL TECH"
c.font = TITLE_FONT
c.alignment = Alignment(horizontal='center')

# Summary stats
stats = [
    ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('Total Leads Collected', len(all_leads)),
    ('Active Boosters (Running Ads)', len(booster_leads)),
    ('Hot Leads (No Website + Phone)', len(hot_leads)),
    ('', ''),
    ('Score Breakdown', ''),
    ('  Score 12 (🔥 HOT)', sum(1 for l in all_leads if l['_score'] == 12)),
    ('  Score 10-11', sum(1 for l in all_leads if 10 <= l['_score'] <= 11)),
    ('  Score 8-9', sum(1 for l in all_leads if 8 <= l['_score'] <= 9)),
    ('  Score < 8', sum(1 for l in all_leads if l['_score'] < 8)),
    ('', ''),
    ('Ad Status', ''),
    ('  Currently Running Ads', len(booster_leads)),
    ('  Not Running / Unknown', len(all_leads) - len(booster_leads)),
    ('', ''),
    ('Website Status', ''),
    ('  Has Website', sum(1 for l in all_leads if l['_has_web'])),
    ('  No Website', sum(1 for l in all_leads if not l['_has_web'])),
]

for ri, (label, val) in enumerate(stats, 3):
    c1 = ws4.cell(row=ri, column=1, value=label)
    c2 = ws4.cell(row=ri, column=2, value=val)
    if label.startswith('  '):
        c1.font = NORMAL
    elif label:
        c1.font = SECTION_FONT
    c2.font = NORMAL if not isinstance(val, int) or val == '' else Font(name='Calibri', bold=True, size=10)

# Hot leads color
for ri in range(3, len(stats) + 3):
    if ws4.cell(row=ri, column=1).value and 'Hot Leads' in str(ws4.cell(row=ri, column=1).value):
        ws4.cell(row=ri, column=2).font = RED_BOLD
    if ws4.cell(row=ri, column=1).value and 'Total Leads' in str(ws4.cell(row=ri, column=1).value):
        ws4.cell(row=ri, column=2).font = Font(name='Calibri', bold=True, size=14, color=NAVY)

# Category breakdown
cat_row = len(stats) + 5
ws4.cell(row=cat_row, column=1, value='Category Breakdown').font = SECTION_FONT
cat_row += 1

cat_counts = Counter(l['_category'] for l in all_leads)
cat_headers = ['Category', 'Total', 'Active Boosters', 'Hot (No Web)', '% of Total']
for ci, h in enumerate(cat_headers, 1):
    c = ws4.cell(row=cat_row, column=ci, value=h)
    c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    c.border = THIN_BORDER
cat_row += 1

for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
    boost_count = sum(1 for l in all_leads if l['_category'] == cat and is_running_ads(l))
    hot_count = sum(1 for l in all_leads if l['_category'] == cat and is_hot(l))
    pct = f"{count/len(all_leads)*100:.1f}%"
    
    ws4.cell(row=cat_row, column=1, value=cat).border = THIN_BORDER
    ws4.cell(row=cat_row, column=2, value=count).border = THIN_BORDER
    c3 = ws4.cell(row=cat_row, column=3, value=boost_count)
    c3.border = THIN_BORDER
    c3.font = Font(color='006100', bold=True) if boost_count > 0 else NORMAL
    c4 = ws4.cell(row=cat_row, column=4, value=hot_count)
    c4.border = THIN_BORDER
    c4.font = Font(color='B22222', bold=True) if hot_count > 0 else NORMAL
    ws4.cell(row=cat_row, column=5, value=pct).border = THIN_BORDER
    cat_row += 1

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 15

# ════════════════════════════════════
# Save
# ════════════════════════════════════
wb.save(OUT)
print(f"\n{'='*60}")
print(f"✅ PROFESSIONAL REPORT SAVED")
print(f"{'='*60}")
print(f"   File: {OUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Total Leads: {len(all_leads)}")
print(f"   Active Boosters: {len(booster_leads)}")
print(f"   Hot Leads (No Website): {len(hot_leads)}")
print(f"   Categories: {len(cat_counts)}")
